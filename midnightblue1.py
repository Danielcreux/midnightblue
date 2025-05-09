import os
import json
import smtplib
import time
import openpyxl
import threading
import ttkbootstrap as ttk
from tkinter import filedialog, StringVar, messagebox, scrolledtext, Listbox
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formatdate
from PIL import Image, ImageTk  # Importar Pillow para trabajar con imágenes
import webbrowser
import datetime
from datetime import datetime
import csv
from string import Template
import os.path

# Archivo de configuración
CONFIG_FILE = "config.json"

def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r") as f:
            return json.load(f)
    return {}

def save_config():
    config = {
        "smtp_server": smtp_var.get(),
        "username": user_var.get(),
        "password": pass_var.get()
    }
    with open(CONFIG_FILE, "w") as f:
        json.dump(config, f)

def select_html():
    file_path = filedialog.askopenfilename(filetypes=[["HTML files", "*.html"]])
    html_var.set(file_path)

def select_excel():
    file_path = filedialog.askopenfilename(filetypes=[["Excel files", "*.xlsx;*.ods"]])
    excel_var.set(file_path)

def log_message(message):
    timestamp = time.strftime("%H:%M:%S")
    formatted_message = f"[{timestamp}] {message}"
    console_text.config(state='normal')
    console_text.insert('end', formatted_message + "\n")
    console_text.config(state='disabled')
    console_text.see('end')  # Asegura que siempre se vea el último mensaje



def preview_html():
    html_file = html_var.get()
    if not html_file:
        messagebox.showerror("Error", "Por favor seleccione un archivo HTML primero")
        return
    
    try:
        webbrowser.open(html_file)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo abrir el archivo HTML: {e}")

def schedule_email():
    def save_schedule():
        scheduled_time = datetime.strptime(f"{date_var.get()} {time_var.get()}", "%Y-%m-%d %H:%M")
        if scheduled_time <= datetime.now():
            messagebox.showerror("Error", "La fecha y hora deben ser futuras")
            return
        
        schedule_window.destroy()
        # Programar el envío
        threading.Timer((scheduled_time - datetime.now()).total_seconds(), send_emails).start()
        log_message(f"✓ Envío programado para: {scheduled_time}")

    schedule_window = ttk.Toplevel(root)
    schedule_window.title("Programar Envío")
    schedule_window.geometry("300x200")
    
    date_var = StringVar(value=datetime.now().strftime("%Y-%m-%d"))
    time_var = StringVar(value=datetime.now().strftime("%H:%M"))
    
    ttk.Label(schedule_window, text="Fecha (YYYY-MM-DD):").pack(pady=5)
    ttk.Entry(schedule_window, textvariable=date_var).pack(fill='x', padx=20)
    
    ttk.Label(schedule_window, text="Hora (HH:MM):").pack(pady=5)
    ttk.Entry(schedule_window, textvariable=time_var).pack(fill='x', padx=20)
    
    ttk.Button(schedule_window, text="Programar", command=save_schedule).pack(pady=20)

def save_template():
    html_file = html_var.get()
    if not html_file:
        messagebox.showerror("Error", "Seleccione un archivo HTML primero")
        return
    
    template_name = template_var.get()
    if not template_name:
        messagebox.showerror("Error", "Ingrese un nombre para la plantilla")
        return
    
    templates_dir = "templates"
    if not os.path.exists(templates_dir):
        os.makedirs(templates_dir)
    
    template_path = os.path.join(templates_dir, f"{template_name}.html")
    try:
        with open(html_file, 'r', encoding='utf-8') as source:
            with open(template_path, 'w', encoding='utf-8') as target:
                target.write(source.read())
        log_message(f"✓ Plantilla guardada como: {template_name}")
    except Exception as e:
        messagebox.showerror("Error", f"Error al guardar la plantilla: {e}")

def load_template():
    templates_dir = "templates"
    if not os.path.exists(templates_dir):
        messagebox.showerror("Error", "No hay plantillas guardadas")
        return
    
    template_files = [f for f in os.listdir(templates_dir) if f.endswith('.html')]
    if not template_files:
        messagebox.showerror("Error", "No hay plantillas disponibles")
        return
    
    def select_template():
        selected = template_listbox.get(template_listbox.curselection())
        html_var.set(os.path.join(templates_dir, selected))
        template_window.destroy()
    
    template_window = ttk.Toplevel(root)
    template_window.title("Cargar Plantilla")
    template_window.geometry("300x400")
    
    template_listbox = Listbox(template_window)
    template_listbox.pack(fill='both', expand=True, padx=20, pady=20)
    
    for template in template_files:
        template_listbox.insert('end', template)
    
    ttk.Button(template_window, text="Seleccionar", command=select_template).pack(pady=10)


generate_report_stats = []

def generate_report():
    global generate_report_stats
    if not generate_report_stats:
        generate_report_stats = []
    
    report_file = "reporte_envios.csv"
    with open(report_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['Fecha', 'Destinatario', 'Estado', 'Error'])
        writer.writerows(generate_report_stats)
    
    log_message(f"✓ Reporte generado: {report_file}")
    webbrowser.open(report_file)

def send_emails():
    def task():
        global generate_report_stats
        smtp_server = smtp_var.get()
        username = user_var.get()
        password = pass_var.get()
        html_file = html_var.get()
        excel_file = excel_var.get()
        
        if not all([smtp_server, username, password, html_file, excel_file]):
            messagebox.showerror("Error", "Todos los campos son obligatorios")
            return
        
        try:
            with open(html_file, "r", encoding="utf-8") as file:
                html_content = file.read()
        except FileNotFoundError:
            messagebox.showerror("Error", "Archivo HTML no encontrado")
            return
        
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        total_emails = sheet.max_row - 1
        root.after(0, lambda: progress_var.set(0))
        
        try:
            server = smtplib.SMTP(smtp_server, 587)
            server.starttls()
            server.login(username, password)
            log_message("Conexión exitosa al servidor SMTP")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo conectar al servidor SMTP: {e}")
            return
        
        sent_count = 0
        total_count = 0
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            receiver_email = row[0]
            if receiver_email:
                # Personalización del mensaje
                personalization = {
                    'nombre': row[1] if len(row) > 1 else '',
                    'empresa': row[2] if len(row) > 2 else '',
                    'cargo': row[3] if len(row) > 3 else ''
                }
                
                # Reemplazar placeholders en el contenido HTML
                template = Template(html_content)
                personalized_content = template.safe_substitute(personalization)
                
                message = MIMEMultipart("alternative")
                message["Subject"] = subject_var.get() or "Sin asunto"
                message["From"] = username
                message["To"] = receiver_email
                message["Date"] = formatdate(localtime=True)
                message.attach(MIMEText(personalized_content, "html"))
                
                try:
                    server.sendmail(username, receiver_email, message.as_string())
                    status = "Enviado"
                    error = ""
                except Exception as e:
                    status = "Error"
                    error = str(e)
                
                # Guardar estadísticas
                generate_report_stats.append([
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    receiver_email,
                    status,
                    error
                ])
                
                total_count += 1
                log_message(f"📧 Intentando enviar correo a: {receiver_email}")
                message = MIMEMultipart("alternative")
                message["Subject"] = "Your email subject"
                message["From"] = username
                message["To"] = receiver_email
                message["Date"] = formatdate(localtime=True)
                message.attach(MIMEText(html_content, "html"))
                
                try:
                    server.sendmail(username, receiver_email, message.as_string())
                    sent_count += 1
                    root.after(0, lambda count=sent_count: progress_var.set((count / total_emails) * 100))
                    log_message(f"✅ Correo enviado exitosamente a: {receiver_email}")
                    time.sleep(10)
                except Exception as e:
                    log_message(f"❌ Error al enviar correo a {receiver_email}: {str(e)}")
        
        server.quit()
        root.after(0, lambda: progress_var.set(100))
        log_message(f"✨ Proceso completado. {sent_count} de {total_count} correos enviados exitosamente.")
        messagebox.showinfo("Éxito", f"Proceso completado.\nCorreos enviados: {sent_count}/{total_count}")
    
    thread = threading.Thread(target=task)
    thread.start()

# Cargar configuración previa
config = load_config()

# Crear ventana
root = ttk.Window(themename="superhero")
root.title("Email Sender")
root.geometry("500x550")

# Agregar el logo a la ventana
logo_path = "midnightblue.png"  # Ruta del logo PNG
if os.path.exists(logo_path):
    logo_image = Image.open(logo_path)
    logo_image = logo_image.resize((100, 100), Image.Resampling.LANCZOS)  # Ajustar el tamaño
    logo_photo = ImageTk.PhotoImage(logo_image)
    ttk.Label(root, image=logo_photo).pack(pady=10)  # Mostrar el logo en la ventana

# Variables
smtp_var = StringVar(value=config.get("smtp_server", ""))
user_var = StringVar(value=config.get("username", ""))
pass_var = StringVar(value=config.get("password", ""))
html_var = StringVar()
excel_var = StringVar()
subject_var = StringVar()  # Variable para el asunto
progress_var = ttk.DoubleVar()
template_var = StringVar()  # Moved this up with other variables

# Widgets
ttk.Label(root, text="Asunto del Correo:").pack(pady=5)
ttk.Entry(root, textvariable=subject_var).pack(fill='x', padx=20)
ttk.Label(root, text="SMTP Server:").pack(pady=5)
ttk.Entry(root, textvariable=smtp_var).pack(fill='x', padx=20)

ttk.Label(root, text="Usuario:").pack(pady=5)
ttk.Entry(root, textvariable=user_var).pack(fill='x', padx=20)

ttk.Label(root, text="Contraseña:").pack(pady=5)
ttk.Entry(root, textvariable=pass_var, show='*').pack(fill='x', padx=20)

ttk.Button(root, text="Seleccionar HTML", command=select_html).pack(pady=5)
ttk.Entry(root, textvariable=html_var, state='readonly').pack(fill='x', padx=20)

ttk.Button(root, text="Seleccionar Excel", command=select_excel).pack(pady=5)
ttk.Entry(root, textvariable=excel_var, state='readonly').pack(fill='x', padx=20)

# Template widgets
ttk.Label(root, text="Nombre de Plantilla:").pack(pady=5)
ttk.Entry(root, textvariable=template_var).pack(fill='x', padx=20)

# Frame para botones de gestión
management_frame = ttk.Frame(root)
management_frame.pack(fill='x', padx=20, pady=10)

ttk.Button(management_frame, text="Vista Previa", command=preview_html).pack(side='left', padx=5)
ttk.Button(management_frame, text="Programar Envío", command=schedule_email).pack(side='left', padx=5)
ttk.Button(management_frame, text="Guardar Plantilla", command=save_template).pack(side='left', padx=5)
ttk.Button(management_frame, text="Cargar Plantilla", command=load_template).pack(side='left', padx=5)
ttk.Button(management_frame, text="Generar Reporte", command=generate_report).pack(side='left', padx=5)

ttk.Button(root, text="Enviar Correos", command=send_emails, bootstyle='success').pack(pady=20)
ttk.Progressbar(root, variable=progress_var, maximum=100).pack(fill='x', padx=20, pady=5)

ttk.Label(root, text="Consola de mensajes:").pack(pady=5)
console_text = scrolledtext.ScrolledText(root, height=12, state='disabled')
console_text.pack(fill='both', padx=20, pady=5)

ttk.Button(root, text="Guardar Configuración", command=save_config, bootstyle='info').pack(pady=5)

# Ejecutar ventana (moved to the end)
root.mainloop()


