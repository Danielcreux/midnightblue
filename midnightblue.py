import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formatdate
import time
import openpyxl

# Configuración del servidor SMTP y credenciales
smtp_server = "smtp.ionos.es"
port = 587
sender_email = "pruebas@freire-sanchez-valencia.es"
sender_password = "Prueba_123@"  # Usa variables de entorno en producción

# Archivo HTML con la plantilla del correo
html_file = "email_template.html"

# Leer el contenido del archivo HTML
try:
    with open(html_file, "r", encoding="utf-8") as file:
        html_content = file.read()
except FileNotFoundError:
    print("Error: El archivo HTML no se encontró.")
    exit()

# Leer direcciones de correo desde el archivo Excel
excel_file = 'emails.xlsx'
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active

# Se asume que los correos están en la primera columna
email_column = 1

# Conectar al servidor SMTP solo una vez para evitar múltiples autenticaciones
try:
    server = smtplib.SMTP(smtp_server, port)
    server.ehlo()
    server.starttls()  # Seguridad
    server.ehlo()
    server.login(sender_email, sender_password)
    print("Conexión exitosa al servidor SMTP")

    # Enviar correos individualmente
    for row in sheet.iter_rows(min_row=2, max_col=email_column, max_row=sheet.max_row):
        receiver_email = row[0].value  # Obtener email de la primera columna

        if receiver_email:
            # Crear un nuevo objeto de correo en cada iteración
            message = MIMEMultipart("alternative")
            message["Subject"] = "Your email subject"
            message["From"] = sender_email
            message["To"] = receiver_email
            message["Date"] = formatdate(localtime=True)

            # Adjuntar el contenido HTML
            html_part = MIMEText(html_content, "html")
            message.attach(html_part)

            try:
                # Enviar correo
                server.sendmail(sender_email, receiver_email, message.as_string())
                print(f"✅ Email enviado a {receiver_email}")

                # Pausa para evitar bloqueos por parte de IONOS
                time.sleep(10)

            except Exception as e:
                print(f"❌ Error enviando email a {receiver_email}: {e}")

    print("✅ Todos los correos han sido enviados.")

except Exception as e:
    print(f"❌ Error al conectar con el servidor SMTP: {e}")

finally:
    server.quit()  # Cerrar la conexión SMTP
    print("Conexión SMTP cerrada.")
